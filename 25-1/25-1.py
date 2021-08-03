from openpyxl import load_workbook


wb = load_workbook('email_list.xlsx', read_only=True)
data = wb.active
# for row in data.iter_rows(max_col=3):
#   for cell in row :
#     if not cell.value=='1':
#       print(cell.value)

  # if row[0].value=='NONE':
  #   break

for row in data[3:data.max_row]:
    
       print(row[1].value)
       print(row[2].value)